from openpyxl.styles import Border, Side

def init(wb, sheet_name):
    # 엑셀 파일에 시트 생성
    ws = wb.create_sheet(sheet_name)

def edit_header(wb, sheet_name, header_title_list, header_text_list):
    # 엑셀 파일에서 시트 이름이 sheet_name과 일치하는 시트 지정
    ws = wb.get_sheet_by_name(sheet_name)

    # 시트 양식에 맞게 입력
    write_list_to_excel(ws,2,2,header_title_list[:1])
    write_list_to_excel(ws,2,3,header_title_list[1:]) # write_list_to_excel(ws,2,4,header_title_list[1:])

    # 시트 양식에 맞게 입력
    write_list_to_excel(ws,3,2,header_text_list[:1])
    write_list_to_excel(ws,3,3,header_text_list[1:])

    # 시트 양식에 맞게 테두리 지정
    add_border_to_range(ws, 2, 2, 3, 2) # 테두리를 가로로 입력
    add_border_to_range(ws, 2, 3, 3, 1 + len(header_title_list)) #  add_border_to_range(ws, 2, 3, 3, 2 + len(header_title_list)), 1은 한칸 띄우고 시작하니까.

def edit_content(wb, sheet_name, content_top_title_list, content_sub_title_list, content_text_list):
    # 엑셀 파일에서 시트 이름이 sheet_name과 일치하는 시트 지정
    ws = wb.get_sheet_by_name(sheet_name)

    # 시트 양식에 맞게 입력
    write_list_to_excel(ws,5,4,content_top_title_list[:2],5) # 국어 영역, 수학 영역
    write_list_to_excel(ws,5,14,content_top_title_list[2:4]) # 영어 영역, 한국사 영역
    
    # 시트 양식에 맞게 셀 병합
    ws.merge_cells("D5:H5") # 국어 영역
    ws.merge_cells("I5:M5") # 수학 영역

    if content_top_title_list.count('제2외국어/한문 영역'):
        # 시트 양식에 맞게 입력
        write_list_to_excel(ws,5,16,content_top_title_list[4:],10) # 탐구영역

        # 시트 양식에 맞게 셀 병합
        ws.merge_cells("P5:Y5")
        ws.merge_cells("Z5:AA5")

    else:
        # 시트 양식에 맞게 입력
        write_list_to_excel(ws,5,16,content_top_title_list[4:])

        # 시트 양식에 맞게 셀 병합
        ws.merge_cells("P5:Y5")

    # 시트 양식에 맞게 입력
    write_list_to_excel(ws,6,2,content_sub_title_list[:1],2) # 번호
    write_list_to_excel(ws,6,3,content_sub_title_list[1:]) # 성명

    # 시트 양식에 맞게 입력
    for i in range(len(content_text_list)):
        write_list_to_excel(ws,7+i,2,content_text_list[i][:2]) # 개별 번호, 성명이어야 하는데...? gap = 1이면 됨. 빈 컬럼 삭제
        write_list_to_excel(ws,7+i,4,content_text_list[i][2:]) # 국어 선택과목

    # 시트 양식에 맞게 테두리 지정
    add_border_to_range(ws, 5, 2, 5 + len(content_text_list) + 1, 3) # 번호, 성명 부분 테두리 설정
    add_border_to_range(ws, 5, 4, 5 + len(content_text_list) + 1, 4 + len(content_text_list[0]) - 3) #-3: 4번째 컬럼부터 테두리를 설정해야 하니, 앞에 3개는 패스.

def edit_content_sn(wb, sheet_name, content_top_title_list, content_sub_title_list, content_text_list):
    # 엑셀 파일에서 시트 이름이 sheet_name과 일치하는 시트 지정
    ws = wb.get_sheet_by_name(sheet_name)

    # 시트 양식에 맞게 입력
    write_list_to_excel(ws,5,5,content_top_title_list[0:1]) 
    write_list_to_excel(ws,5,6,content_top_title_list[1:4], 4)
    write_list_to_excel(ws,5,15,content_top_title_list[4:], 8)
    
    # 시트 양식에 맞게 셀 병합
    ws.merge_cells("F5:I5")
    ws.merge_cells("J5:M5")
    ws.merge_cells("O5:V5")

    # 시트 양식에 맞게 입력
    write_list_to_excel(ws,6,2,content_sub_title_list[:1],2)
    write_list_to_excel(ws,6,4,content_sub_title_list[1:])

    # 시트 양식에 맞게 입력
    for i in range(len(content_text_list)):
        write_list_to_excel(ws,7+i,2,content_text_list[i][:2],2)
        write_list_to_excel(ws,7+i,5,content_text_list[i][2:])

    # 시트 양식에 맞게 테두리 지정
    add_border_to_range(ws, 5, 2, 5 + len(content_text_list) + 1, 2)
    add_border_to_range(ws, 5, 4, 5 + len(content_text_list) + 1, 4 + len(content_text_list[0]) - 2)

# 미사용 주석처리
# def edit_footer(wb, sheet_name, list):
#     print("here3")

def write_list_to_excel(ws, start_row, start_col, input_list, gap = 1):
    # 리스트 값들을 차례로 입력
    current_row = start_row
    current_col = start_col
    
    for value in input_list:
        ws.cell(row=current_row, column=current_col, value=value)
        current_col += gap

def add_border_to_range(ws, start_row, start_col, end_row, end_col):
    # Border 객체 생성
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # 특정 범위의 셀에 테두리 설정
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border