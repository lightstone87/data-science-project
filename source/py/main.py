import sys

from openpyxl import Workbook

from source.py import pdf_to_dataframe as p2df
from source.py import convert
from source.py import edit

# def conv_and_edit(wb, df_list):

def load_file_1st(parent = None, file_path = "./", folder_path = "./result/dist/xlsx/"):
    try:
        # 변환버튼 텍스트 수정
        parent.convert_file_button_text.write("로딩중...")

        # PDF의 표를 DataFrame으로 변환
        # 페이지 순서대로 list에 저장
        df_list = p2df.get_df_list_1st(file_path)
        
        # 엑셀 새로운 파일 만들기 위해 생성
        wb = Workbook()

        # PDF 파일의 페이지 수 반복
        for page_df_list in df_list:
            # page_df_list는 header dataframe과 content dataframe로 구성
            # page_df_list[0]은 header, page_df_list[1]은 content에 해당하는 내용

            # header_title_list는 header칸의 타이틀 정보를 포함한 list
            # header_text_list는 header칸의 내용 정보를 포함한 list
            header_title_list, header_text_list = convert.conv_header_1st(page_df_list[0])

            # content_top_title_list는 content칸의 영역 정보를 포함한 list
            # content_sub_title_list는 content칸의 성적 타이틀 정보를 포함한 list
            # content_text_list는 content칸의 학생별 정보를 포함한 list
            content_top_title_list, content_sub_title_list, content_text_list = convert.conv_content_1st(page_df_list[1])

            # 엑셀 시트 생성시 필요한 시트 이름
            sheet_name = f"{header_text_list[0]}-{header_text_list[1]}"
            parent.convert_file_button_text.write(f"변환중...{sheet_name}")
            
            # 엑셀의 시트 생성 및 기본 양식 작성
            edit.init(wb, sheet_name)
            
            # 시트의 header에 해당하는 부분 작성
            edit.edit_header(wb, sheet_name, header_title_list, header_text_list)
            
            # 시트의 content에 해당하는 부분 작성
            edit.edit_content_1st(wb, sheet_name, content_top_title_list, content_sub_title_list, content_text_list)

        # 엑셀 파일의 시트 이름 list 확인
        ws_name_list = wb.sheetnames
        
        # 엑셀 새로운 파일 생성시 시트 이름이 "Sheet" 시트가 생성
        # 불필요한 시트 제거를 위한 작업
        if "Sheet" in ws_name_list:
            wb.remove(wb["Sheet"])
        
        # 파일의 시트 이름 list 확인
        ws_name_list = wb.sheetnames

        # 엑셀 파일 이름을 지정
        date = header_text_list[-1].replace('.','')
        date = date.replace(' ','')
        grade = header_text_list[0]
        class_num = header_text_list[1]

        file_name = f"{folder_path}/{date}_{grade}학년_{class_num}반_성적표.xlsx"

        parent.convert_file_button_text.write(f"저장중...")
        # 엑셀 파일 저장
        wb.save(file_name)

        # 변환 결과를 확인하기 위한 작업
        if parent:
            parent.convert_file_button_text.write("변환 완료")
        else:
            print("saved")
    except:
        # 변환 결과를 확인하기 위한 작업
        if parent:
            parent.convert_file_button_text.write("변환 실패")
        else:
            print("error")

def load_file(parent = None, file_path = "./", folder_path = "./result/dist/xlsx/"):
    try:
        # 변환버튼 텍스트 수정
        parent.convert_file_button_text.write("로딩중...")

        # PDF의 표를 DataFrame으로 변환
        # 페이지 순서대로 list에 저장
        df_list = p2df.get_df_list(file_path)
        
        # 엑셀 새로운 파일 만들기 위해 생성
        wb = Workbook()

        # PDF 파일의 페이지 수 반복
        for page_df_list in df_list:
            # page_df_list는 header dataframe과 content dataframe로 구성
            # page_df_list[0]은 header, page_df_list[1]은 content에 해당하는 내용

            # header_title_list는 header칸의 타이틀 정보를 포함한 list
            # header_text_list는 header칸의 내용 정보를 포함한 list
            header_title_list, header_text_list = convert.conv_header(page_df_list[0])

            # content_top_title_list는 content칸의 영역 정보를 포함한 list
            # content_sub_title_list는 content칸의 성적 타이틀 정보를 포함한 list
            # content_text_list는 content칸의 학생별 정보를 포함한 list
            content_top_title_list, content_sub_title_list, content_text_list = convert.conv_content(page_df_list[1])
            
            # 엑셀 시트 생성시 필요한 시트 이름
            sheet_name = f"{header_text_list[0]}-{header_text_list[1]}"
            parent.convert_file_button_text.write(f"변환중...{sheet_name}")
            
            # 엑셀의 시트 생성 및 기본 양식 작성
            edit.init(wb, sheet_name)
            
            # 시트의 header에 해당하는 부분 작성
            edit.edit_header(wb, sheet_name, header_title_list, header_text_list)
            
            # 시트의 content에 해당하는 부분 작성
            edit.edit_content(wb, sheet_name, content_top_title_list, content_sub_title_list, content_text_list)

        # 엑셀 파일의 시트 이름 list 확인
        ws_name_list = wb.sheetnames
        
        # 엑셀 새로운 파일 생성시 시트 이름이 "Sheet" 시트가 생성
        # 불필요한 시트 제거를 위한 작업
        if "Sheet" in ws_name_list:
            wb.remove(wb["Sheet"])
        
        # 파일의 시트 이름 list 확인
        ws_name_list = wb.sheetnames

        # 엑셀 파일 이름을 지정
        date = header_text_list[-1].replace('.','')
        date = date.replace(' ','')
        grade = header_text_list[0]

        file_name = f"{folder_path}/{date}_{grade}학년_성적표.xlsx"

        parent.convert_file_button_text.write(f"저장중...")
        # 엑셀 파일 저장
        wb.save(file_name)

        # 변환 결과를 확인하기 위한 작업
        if parent:
            parent.convert_file_button_text.write("변환 완료")
        else:
            print("saved")
    except:
        # 변환 결과를 확인하기 위한 작업
        if parent:
            parent.convert_file_button_text.write("변환 실패")
        else:
            print("error")

def load_file_sn(parent = None, file_path = "./", folder_path = "./result/dist/xlsx/"):
    try:
        # 변환버튼 텍스트 수정
        parent.convert_file_button_text.write("로딩중...")

        # PDF의 표를 DataFrame으로 변환
        # 페이지 순서대로 list에 저장
        df_list = p2df.get_df_list_sn(file_path)

        file_name = df_list[0][0]
        school_name = df_list[0][1]
        class_num = df_list[0][2]

        # 엑셀 새로운 파일 만들기 위해 생성
        wb = Workbook()

        # 엑셀의 시트 생성 및 기본 양식 작성
        edit.init(wb, class_num)

        # header_text_list는 header칸의 내용 정보를 포함한 list
        header_title_list, header_text_list = convert.conv_header_sn(class_num, school_name)

        # 시트의 header에 해당하는 부분 작성
        edit.edit_header(wb, class_num, header_title_list, header_text_list)
        
        content_text_list_total = []

        # PDF 파일의 페이지 수 반복
        for page_df in df_list[1:]:
            # page_df_list는 header dataframe과 content dataframe로 구성
            # page_df_list[0]은 header, page_df_list[1]은 content에 해당하는 내용

            # content_top_title_list는 content칸의 영역 정보를 포함한 list
            # content_sub_title_list는 content칸의 성적 타이틀 정보를 포함한 list
            # content_text_list는 content칸의 학생별 정보를 포함한 list
            content_top_title_list, content_sub_title_list, content_text_list = convert.conv_content_sn(page_df)
            content_text_list_total += content_text_list
            # print(content_top_title_list)

            # # 엑셀 시트 생성시 필요한 시트 이름
            # sheet_name = f"{header_text_list[0]}-{header_text_list[1]}"
            parent.convert_file_button_text.write(f"변환중...")
            
            
        # # 시트의 content에 해당하는 부분 작성
        edit.edit_content_sn(wb, class_num, content_top_title_list, content_sub_title_list, content_text_list_total)

        # 엑셀 파일의 시트 이름 list 확인
        ws_name_list = wb.sheetnames
        
        # 엑셀 새로운 파일 생성시 시트 이름이 "Sheet" 시트가 생성
        # 불필요한 시트 제거를 위한 작업
        if "Sheet" in ws_name_list:
            wb.remove(wb["Sheet"])
        
        # 파일의 시트 이름 list 확인
        ws_name_list = wb.sheetnames

        parent.convert_file_button_text.write(f"저장중...")
        # 엑셀 파일 저장
        wb.save(f"{folder_path}/{file_name}.xlsx")

        # 변환 결과를 확인하기 위한 작업
        if parent:
            parent.convert_file_button_text.write("변환 완료")
        else:
            print("saved")
    except:
        # 변환 결과를 확인하기 위한 작업
        if parent:
            parent.convert_file_button_text.write("변환 실패")
        else:
            print("error")

if __name__ == '__main__':
    load_file(None, sys.argv[1])